from openpyxl import load_workbook, Workbook
from datetime import datetime
import re


class Controller:
    
    def __init__(self) -> None:
        self.filename_format = '%Y-%m-%d'
        self.new_wb()
    
    def filename(self):
        return f"{datetime.now().strftime(self.filename_format)}.xlsx"
    
    def new_wb(self):
        wb=''
        self._filename = self.filename()
        try:
            file_day = datetime.strptime(self._filename.split('.')[0], self.filename_format).day
            if (datetime.now().day > file_day):
                raise FileNotFoundError()
            wb=load_workbook(self._filename, read_only=False)
        except FileNotFoundError:
            wb=Workbook()
            wb.active['A1'].value=''
            wb.save(filename = self._filename)
        self.wb=wb
        self.sheet= self.wb.active
    
    def __save_to_excel(self, data:list):
        self.sheet.append(data)
        self.wb.save(self._filename)
        self.new_wb()
    
    def col_4(self):
        d= [c.value for c in self.sheet['D']]
        self.new_wb()
        return d

    def get_excel_rows(self):
        self.new_wb()
        excel_data=[]
        for row in self.sheet.rows:
            if row[0].value is not None:
                excel_data.append([row[0].value, row[1].value, row[2].value, row[3].value, row[4].value])
        excel_data.reverse()
        return excel_data
    
    def get_total_count(self):
        max_row = self.sheet.max_row - 1
        self.new_wb()
        return max_row

    def check_contents(self, content):
        if re.match("^[0-9]+$",content):
            duplicate = ''
            if content[2:] in self.col_4():
                duplicate = '1'
            self.__save_to_excel([content,datetime.now().strftime('%Y-%m-%d %H:%M:%S'), content[:2], content[2:], duplicate]) 
            return True 
        else:
            return False